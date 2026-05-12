"""§10.10 — Output generation: XLSX (4 sheets) and JSON."""

import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from invoice_parser.schema.invoice import Invoice
from invoice_parser.worker.pipeline import InvoiceResult
from invoice_parser.worker.segment import SkippedPage

_GREEN = PatternFill("solid", fgColor="C6EFCE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_RED = PatternFill("solid", fgColor="FFC7CE")
_HEADER_FONT = Font(bold=True)

_INVOICE_COLS = [
    "id", "source_file", "source_pages", "invoice_index_in_file", "status", "notes",
    "vendor_name", "vendor_address", "vendor_tax_id", "vendor_email", "vendor_phone",
    "bill_to_name", "bill_to_address", "invoice_number", "invoice_date", "due_date",
    "po_number", "currency", "subtotal", "tax_amount", "tax_rate", "discount",
    "shipping", "total", "amount_due", "cost_usd", "models_used",
]

_AUDIT_COLS = [
    "invoice_id", "field_name", "final_value", "decision",
    "ocr_value", "ocr_confidence", "ocr_token_id", "ocr_bbox",
    "vision_value", "vision_confidence",
]


def _autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)


def _write_header(ws, cols: list[str]):
    ws.append(cols)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"


def _inv_row(inv_id: str, inv: Invoice, result: InvoiceResult) -> list:
    return [
        inv_id,
        inv.source_file,
        str(inv.source_pages),
        inv.invoice_index_in_file,
        inv.status,
        inv.notes,
        inv.vendor_name,
        inv.vendor_address,
        inv.vendor_tax_id,
        inv.vendor_email,
        inv.vendor_phone,
        inv.bill_to_name,
        inv.bill_to_address,
        inv.invoice_number,
        str(inv.invoice_date) if inv.invoice_date else None,
        str(inv.due_date) if inv.due_date else None,
        inv.po_number,
        inv.currency,
        float(inv.subtotal) if inv.subtotal is not None else None,
        float(inv.tax_amount) if inv.tax_amount is not None else None,
        float(inv.tax_rate) if inv.tax_rate is not None else None,
        float(inv.discount) if inv.discount is not None else None,
        float(inv.shipping) if inv.shipping is not None else None,
        float(inv.total) if inv.total is not None else None,
        float(inv.amount_due) if inv.amount_due is not None else None,
        round(result.cost_usd, 6),
        ", ".join(result.models_used),
    ]


def generate_xlsx(
    job_id: str,
    results: list[tuple[str, InvoiceResult]],  # (inv_id, result)
    skipped_pages: list[SkippedPage],
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Invoices ────────────────────────────────────────────────────
    ws_inv = wb.active
    ws_inv.title = "Invoices"
    _write_header(ws_inv, _INVOICE_COLS)
    fill_map = {"ok": _GREEN, "review": _YELLOW, "failed": _RED}
    for inv_id, result in results:
        inv = result.invoice
        row = _inv_row(inv_id, inv, result)
        ws_inv.append(row)
        fill = fill_map.get(inv.status)
        if fill:
            status_col = _INVOICE_COLS.index("status") + 1
            ws_inv.cell(ws_inv.max_row, status_col).fill = fill
    _autosize(ws_inv)

    # ── Sheet 2: Line Items ──────────────────────────────────────────────────
    ws_li = wb.create_sheet("Line Items")
    _write_header(ws_li, ["invoice_id", "invoice_number", "line_number",
                           "description", "quantity", "unit_price", "line_total", "confidence"])
    for inv_id, result in results:
        inv = result.invoice
        for item in inv.line_items:
            ws_li.append([
                inv_id, inv.invoice_number, item.line_number,
                item.description,
                float(item.quantity) if item.quantity is not None else None,
                float(item.unit_price) if item.unit_price is not None else None,
                float(item.line_total) if item.line_total is not None else None,
                item.confidence,
            ])
    _autosize(ws_li)

    # ── Sheet 3: Review ──────────────────────────────────────────────────────
    ws_rev = wb.create_sheet("Review")
    _write_header(ws_rev, _INVOICE_COLS + ["reason"])
    for inv_id, result in results:
        inv = result.invoice
        if inv.status != "ok":
            row = _inv_row(inv_id, inv, result) + [inv.notes]
            ws_rev.append(row)
    _autosize(ws_rev)

    # ── Sheet 4: Audit ───────────────────────────────────────────────────────
    ws_audit = wb.create_sheet("Audit")
    _write_header(ws_audit, _AUDIT_COLS)
    for inv_id, result in results:
        for fname, entry in result.audit.items():
            ocr = entry.get("ocr") or {}
            vis = entry.get("vision_llm") or {}
            ws_audit.append([
                inv_id, fname,
                entry.get("final_value"),
                entry.get("decision"),
                ocr.get("value"),
                ocr.get("confidence"),
                ocr.get("token_id"),
                str(ocr.get("bbox", "")),
                vis.get("value"),
                vis.get("self_confidence"),
            ])
    _autosize(ws_audit)

    # ── Sheet 5: Skipped ─────────────────────────────────────────────────────
    ws_skip = wb.create_sheet("Skipped")
    _write_header(ws_skip, ["page", "reason"])
    for sp in skipped_pages:
        ws_skip.append([sp.page, sp.reason])
    _autosize(ws_skip)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_json(
    job_id: str,
    results: list[tuple[str, InvoiceResult]],
    skipped_pages: list[SkippedPage],
) -> bytes:
    invoices_out = []
    for inv_id, result in results:
        inv = result.invoice
        inv_dict = inv.model_dump(mode="json")
        inv_dict["id"] = inv_id
        inv_dict["audit"] = result.audit
        invoices_out.append(inv_dict)

    summary = {
        "invoices_total": len(results),
        "invoices_ok": sum(1 for _, r in results if r.invoice.status == "ok"),
        "invoices_review": sum(1 for _, r in results if r.invoice.status == "review"),
        "invoices_failed": sum(1 for _, r in results if r.invoice.status == "failed"),
        "total_cost_usd": round(sum(r.cost_usd for _, r in results), 6),
    }

    output = {
        "job_id": job_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": summary,
        "invoices": invoices_out,
        "skipped_pages": [{"page": sp.page, "reason": sp.reason} for sp in skipped_pages],
    }
    return json.dumps(output, indent=2, default=str).encode()
